import re
import openpyxl
from pathlib import Path
import os
import json


def split_workbook(input_file, output_dir):
  """Splits the worksheets in an Excel workbook into separate workbooks.

  Args:
    input_file: The path to the Excel workbook to split.
    output_dir: The directory to write the output workbooks to.
  """

  workbook = openpyxl.load_workbook(input_file)

  for sheet in workbook.worksheets:
    output_file = os.path.join(output_dir, sheet.title + ".xlsx")
    new_workbook = openpyxl.Workbook()
    new_workbook.remove(new_workbook.active)
    new_worksheet = new_workbook.create_sheet(sheet.title)


    for row in sheet.iter_rows():
      for cell in row:
        # print(cell.column)
        new_worksheet.cell(row=cell.row, column=cell.column).value = cell.value

    new_workbook.save(output_file)


# def split_workbook(input_file, output_dir):
#   """Splits the worksheets in an Excel workbook into separate workbooks.
#
#   Args:
#     input_file: The path to the Excel workbook to split.
#     output_dir: The directory to write the output workbooks to.
#   """
#
#   workbook = openpyxl.load_workbook(input_file)
#
#   for sheet in workbook.worksheets:
#     print(sheet.title)
#     output_file = os.path.join(output_dir, sheet.title + ".xlsx")
#     new_workbook = openpyxl.Workbook()
#     new_workbook.create_sheet(sheet.title)
#     new_workbook.copy_worksheet(sheet)
#     new_workbook.save(output_file)

if __name__ == "__main__":
  input_file = "./excel_sheets/Copy of DHIS2 EMR for testing.xlsx"
  output_dir = "./excel_sheets"
  split_workbook(input_file, output_dir)